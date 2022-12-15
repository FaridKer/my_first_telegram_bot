from aiogram import Bot, Dispatcher, executor, types     # для ТГ бота
import requests                                          # для запроса с сайтов
from bs4 import BeautifulSoup                            # для скрапинга
from openpyxl import load_workbook                       # для перевода в excel


# Раздел, связанный с настройкой бота. Как я понял из статьи на хабре -
# нужно задать переменную для своего токена ТГ, после чего передаем боту
# команду инициализации, после чего некий диспатчер, не понимаю значение
API_TOKEN = '5908870637:AAGqVbjT9V6SZb3zWlC2NGiOA4E9pkdB-BE'
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)


# Благодаря строке @dp.message_handler мы задаем функцию, которая запускается,
# когда пользователь пишет "/start". В ответ бот пишет ему автоответ.
# В статье написано, что обязательно надо писать await, так как программа
# работает асинхронно, но у меня нет понимания, что это значит
@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
   await message.reply("Привет!\nЯ первый бот Фарида!\nОтправь мне любое"
                       " название вакансии, а я тебе обязательно сброшу"
                       " excel-таблицу с доступными вакансиями на hh.ru")


# Благодаря строке @dp.message_handler мы задаем функцию, которая запускается
# с аргументом, содержащим любой текст пользователя.
# В ответ возвращается файл
@dp.message_handler(content_types=["text"])
async def send_file(message: types.Document):
    # ЧАСТЬ ПЕРВАЯ - ФОРМУЛИРОВАНИЕ НУЖНОЙ АДРЕСНОЙ СТРОКИ И ЗАПИСЬ В ФАЙЛ
    # я пытался найти готовые словари с расшифровкой процентной кодировки,
    # но безуспешно. Нагуглил, что вроде как есть библиотеки с методами, но
    # я не смог разобраться. В итоге оказалось проще вручную создать словарь
    url_dict = {'а': '%D0%B0', 'б': '%D0%B1', 'в': '%D0%B2', 'г': '%D0%B3',
                'д': '%D0%B4', 'е': '%D0%B5', 'ё': '%D0%B5', 'ж': '%D0%B6',
                'з': '%D0%B7', 'и': '%D0%B8', 'й': '%D0%B9', 'к': '%D0%BA',
                'л': '%D0%BB', 'м': '%D0%BC', 'н': '%D0%BD', 'о': '%D0%BE',
                'п': '%D0%BF', 'р': '%D1%80', 'с': '%D1%81', 'т': '%D1%82',
                'у': '%D1%83', 'ф': '%D1%84', 'х': '%D1%85', 'ц': '%D1%86',
                'ч': '%D1%87', 'ш': '%D1%88', 'щ': '%D1%89', 'ъ': '%D1%8A',
                'ы': '%D1%8B', 'ь': '%D1%8C', 'э': '%D1%8D', 'ю': '%D1%8E',
                'я': '%D1%8F',
                ' ': '+',
                'А': '%D0%90', 'Б': '%D0%91', 'В': '%D0%92', 'Г': '%D0%93',
                'Д': '%D0%94', 'Е': '%D0%95', 'Ё': '%D0%95', 'Ж': '%D0%96',
                'З': '%D0%97', 'И': '%D0%98', 'Й': '%D0%99', 'К': '%D0%9A',
                'Л': '%D0%9B', 'М': '%D0%9C', 'Н': '%D0%9D', 'О': '%D0%9E',
                'П': '%D0%9F', 'Р': '%D0%A0', 'С': '%D0%A1', 'Т': '%D0%A2',
                'У': '%D0%A3', 'Ф': '%D0%A4', 'Х': '%D0%A5', 'Ц': '%D0%A6',
                'Ч': '%D0%A7', 'Ш': '%D0%A8', 'Щ': '%D0%A9', 'Ъ': '%D0%AA',
                'Ы': '%D0%AB', 'Ь': '%D0%AC', 'Э': '%D0%AD', 'Ю': '%D0%AE',
                'Я': '%D0%AF',
                '1': '1', '2': '2', '3': '3', '4': '4', '5': '5', '6': '6',
                '7': '7', '8': '8', '9': '9', '0': '0',
                'a': 'a', 'b': 'b', 'c': 'c', 'd': 'd', 'e': 'e', 'f': 'f',
                'g': 'g', 'h': 'h', 'i': 'i', 'j': 'j', 'k': 'k', 'l': 'l',
                'm': 'm', 'n': 'n', 'o': 'o', 'p': 'p', 'q': 'q', 'r': 'r',
                's': 's', 't': 't', 'u': 'u', 'v': 'v', 'w': 'w', 'x': 'x',
                'y': 'y', 'z': 'z',
                'A': 'A', 'B': 'B', 'C': 'C', 'D': 'D', 'E': 'E', 'F': 'F',
                'G': 'G', 'H': 'H', 'I': 'I', 'J': 'J', 'K': 'K', 'L': 'L',
                'M': 'M', 'N': 'N', 'O': 'O', 'P': 'P', 'Q': 'Q', 'R': 'R',
                'S': 'S', 'T': 'T', 'U': 'U', 'V': 'V', 'W': 'W', 'X': 'X',
                'Y': 'Y', 'Z': 'Z',
                '!': '%21', '#': '%23', '$': '%24', '%': '%25', '&': '%26',
                "'": "%27", '(': '%28', ')': '%29', '*': '%2A', '+': '%2B',
                ',': '%2C', '/': '%2F', ':': '%3A', ';': '%3B', '=': '%3D',
                '?': '%3F', '@': '%40', '[': '%5B', ']': '%5D'}
    # Изначально я делал без headers, но тогда приходил ответ 404.
    # Погуглив, узнал, что нужна некая "шапка", дающая сайту
    # первичную информацию об устройстве, которая запрашивает инфу.
    # Можно сказать, что это фейс-контроль
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                             'AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/108.0.0.0 Safari/537.36'}

    # Изучив адресную строку на сайте ХХ, обратил внимание, что есть
    # константа - 'https://hh.ru/search/vacancy?text=', а мой запрос вакансии
    # отображается в виде процентной кодировки. Сделав заранее словарь
    # конвертации, можно разбить слово побуквенно на список символов 'vacancy',
    # где ключ заменяется на значение ключа. Чтобы взять фразу пользователя
    # в ТГ, как вводное слово, нагуглил метод 'message.text'. Таким образом,
    # создаем переменную, которая является конструктором текста адресной строки
    vacancy = ''.join([url_dict[i] for i in message.text])
    hh_url = 'https://hh.ru/search/vacancy?text=' + vacancy

    # У меня нет полного понимания команды 'requests.get', но это нужно,
    # чтобы получить ответ от сайта, с приветственной шапкой headers
    my_request = requests.get(hh_url, headers=headers)

    # Далее создаем файл 'vacancy_list.html' в режиме записи байтов,
    # куда записываем в виде сухого текста все данные
    # из переменной 'my_request', в кодировке 'utf-8'
    with open('Vacancy_list.html', 'wb') as output_file:
        output_file.write(my_request.text.encode('utf-8'))

    # ЧАСТЬ ВТОРАЯ - РАБОТА С ФАЙЛОМ И СОСТАВЛЕНИЕ СПИСКА ВАКАНСИЙ
    # Далее открываем файл 'vacancy_list.html' в режиме чтения, после чего
    # применяем функцию 'BeautifulSoup', который, как я понял, каким-то
    # образом структурирует данные в файле, что позволяет, в дальнейшем,
    # с помощью поискового метода '.find()' искать подходящие нам строки
    # кода с нужной нам информацией. Поигравшись с методом, я понял, что все
    # данные в супе построены в виде древа. Таким образом, если мы с помощью
    # метода '.find()' найдем нужный нам кусок древа, то, в дальнейшем,
    # можем воспользоваться методом '.find_all()', который позволяет
    # через цикл for перебирать все элементы древа одного этажа.
    # Заранее необходимо создать пустой список, куда мы будем заносить
    # нужные нам элементы при переборе элементов через цикл for
    with open('Vacancy_list.html', 'r', encoding='utf-8') as vacancy_page:
        soup = BeautifulSoup(vacancy_page, features="lxml")
        vacancy_tab = soup.find('div', {'class': 'vacancy-serp-content'})
        vacancies = vacancy_tab.find_all('div', {'class': 'serp-item'})
        vacancy_list = []
        for vacancy in vacancies:
            # Методика перебора следующая - мы находим нужный нам тэг и
            # прежде чем записывать информацию через метод '.get()', необходимо
            # провести проверку на тип данных, найденных через метод '.find()'.
            # Если тип данных None, это говорит о том, что требуемая нам строка
            # отсутствует, что привело бы к ошибке при написании '.get()'.
            # Чтобы этого избежать, задаем этой переменной значение прочерка,
            # после чего спокойно заносим в список. К слову, нам нужен список с
            # одинаковым числом элементов при каждом цикле for, т.к, в
            # дальнейшем, буду ссылаться на элементы по индексам от [0] до [6]
            vacancy_info = []

            # Строки, используемые при использовании метода '.find()', были
            # мною выведены после анализа исходного кода страницы вакансий.
            # Подобрал их по закономерностям.
            # Отдельно упомяну то, что при первых прогонах программы замечал,
            # что возникают артефакты в виде текста '\u202f', '\xa0'.
            # Как я понял, это табуляция, поэтому избавляюсь от них
            # с помощью метода ''.replace()
            vacancy_link = vacancy.find('a')
            if vacancy_link is not None:
                vacancy_link = vacancy.find('a').get('href')
            else:
                vacancy_link = '-'

            vacancy_name = vacancy.find('a')
            if vacancy_name is not None:
                vacancy_name = vacancy.find('a').text
            else:
                vacancy_name = '-'

            vacancy_salary = vacancy.find('span', {
                'class': 'bloko-header-section-3'})
            if vacancy_salary is not None:
                vacancy_salary = vacancy.find('span', {
                    'class': 'bloko-header-section-3'}).text
                s1 = vacancy_salary.replace('\u202f', ' ')
                s2 = s1.replace('\xa0', ' ')
                vacancy_salary = s2
            else:
                vacancy_salary = '-'

            vacancy_company = vacancy.find('a', {
                'class': 'bloko-link bloko-link_kind-tertiary'})
            if vacancy_company is not None:
                vacancy_company = vacancy.find('a', {
                    'class': 'bloko-link bloko-link_kind-tertiary'}).text
                s1 = vacancy_company.replace('\xa0', ' ')
                vacancy_company = s1
            else:
                vacancy_company = '-'

            vacancy_city = vacancy.find('div', {
                'data-qa': 'vacancy-serp__vacancy-address'})
            if vacancy_city is not None:
                vacancy_city = vacancy.find('div', {
                    'data-qa': 'vacancy-serp__vacancy-address'}).text
                s1 = vacancy_city.replace('\xa0', ' ')
                vacancy_city = s1
            else:
                vacancy_city = '-'

            vacancy_description = vacancy.find('div', {
                'data-qa': 'vacancy-serp__vacancy_snippet_responsibility'})
            if vacancy_description is not None:
                vacancy_description = vacancy.find('div', {
                    'data-qa': 'vacancy-serp__'
                               'vacancy_snippet_responsibility'}).text
            else:
                vacancy_description = '-'

            vacancy_requirement = vacancy.find('div', {
                'data-qa': 'vacancy-serp__vacancy_snippet_requirement'})
            if vacancy_requirement is not None:
                vacancy_requirement = vacancy.find('div', {
                    'data-qa': 'vacancy-serp__'
                               'vacancy_snippet_requirement'}).text
            else:
                vacancy_requirement = '-'
            # Каждую переменную заношу в список 'vacancy_info'
            # в строгой последовательности, так как в дальнейшем
            # буду ссылаться на них по индексам. После того как сформируется
            # список 'vacancy_info', я заношу его в заранее созданный список
            # vacancy_list, после чего цикл фор переходит
            # на следующую переменную
            vacancy_info.append(vacancy_link)
            vacancy_info.append(vacancy_name)
            vacancy_info.append(vacancy_salary)
            vacancy_info.append(vacancy_company)
            vacancy_info.append(vacancy_city)
            vacancy_info.append(vacancy_description)
            vacancy_info.append(vacancy_requirement)
            vacancy_list.append(vacancy_info)

        # ЧАСТЬ ТРЕТЬЯ - ЗАНЕСЕНИЕ ДАННЫХ ПО ВАКАНСИЯМ В EXCEL-ТАБЛИЦУ.
        # Подсмотрел видео на Ютубе, где описывается библиотека openpyxl
        # для работы с excel-таблицами. У меня нет понимания, как именно
        # работают load_workbook(), wb[] и аргументы функций, просто методом
        # проб и ошибок разобрался, что надо написать, чтобы работало.
        # Предварительно при каждой итерации стираются все ячейки 100:100
        # благодаря методам '.delete_cols()' и '.delete_rows()'
        fn = 'vacancy_excel_list.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws.delete_cols(1, 100)
        ws.delete_rows(1, 100)
        ws['A1'] = '№'
        ws['B1'] = 'Ссылка на вакансию'
        ws['C1'] = 'Вакансия'
        ws['D1'] = 'Зарплата'
        ws['E1'] = 'Работодатель'
        ws['F1'] = 'Город'
        ws['G1'] = 'Описание Вакансии'
        ws['H1'] = 'Требование по Вакансии'

        # так как мы ведем запись со второй строки, а индексы начинаются
        # с 0, поставил i = 2 для цикла for
        i = 2
        for vacancy in vacancy_list:
            ws['A' + str(i)] = i - 1
            ws['B' + str(i)] = vacancy[0]
            ws['C' + str(i)] = vacancy[1]
            ws['D' + str(i)] = vacancy[2]
            ws['E' + str(i)] = vacancy[3]
            ws['F' + str(i)] = vacancy[4]
            ws['G' + str(i)] = vacancy[5]
            ws['H' + str(i)] = vacancy[6]
            i += 1
        wb.save(fn)
        wb.close()

    # По окончанию работы программы получаем готовую excel-таблицу для
    # отправки пользователю. Для этого мы пишем следующий код со ссылкой
    # на местоположение файла в режиме чтения байтов
    await message.reply_document(open(r'C:\Users\doggf\OneDrive'
                                      r'\Документы\GitHub'
                                      r'\my_first_telegram_bot'
                                      r'\vacancy_excel_list.xlsx', 'rb'))

# Честно ничего не понимаю в коде ниже, просто скопировал со статьи на хабре.
# В дальнейшем вернусь, чтобы разобраться в функции, аргументах и синтаксисе
if __name__ == '__main__':
   executor.start_polling(dp, skip_updates=True)
